'''
Created on Apr 17, 2010

@author: bartmosley

Utility functions:
    xl_to_date
    xlValue

Classes:
    XLSReader
    XLSXReader
    XLdb
    XLOut

Factory functions:
    timeseries_toXLS

'''
import os
import sys

import xlrd
import xlwt

import openpyxl as pyxl

from urllib2 import urlopen, URLError
from datetime import date, timedelta
from tempfile import mkstemp


def get_xlvalue(h_):
    if hasattr(h_.value, "encode"):
        return h_.value.encode()
    else:
        return h_.value
            
def xl_to_date(xdate, _datemode = 1):
    yyyy, mm, dd, h, m, s =xlrd.xldate_as_tuple(xdate.value, _datemode)
    return date(yyyy, mm, dd)

def xlValue(x, datemode=1, hash_comments=1, strip_text=1):
    try:
        if(x.ctype == xlrd.XL_CELL_EMPTY or
           x.ctype == xlrd.XL_CELL_BLANK or
           x.ctype == xlrd.XL_CELL_ERROR):
            return ""
            
        if(x.ctype == xlrd.XL_CELL_TEXT):
            t = x.value.encode()
            if hash_comments and t[0] == '#':
                return None
            else:
                return t.strip() if strip_text else t
                
        if(x.ctype == xlrd.XL_CELL_DATE):
            yyyy, mm, dd, h, m, s = xlrd.xldate_as_tuple(x.value,
                                                         datemode)
            return date(yyyy, mm, dd)
            
        else:
            # that leaves XL_CELL_NUMBER and XL_CELL_BOOLEAN
            return x.value
            
    except:
        return None 

class XLSReader(object):
    def __init__(self, url, localfile=True, hash_comments=True):
        self.book = None
        self.hash_comments = hash_comments
        
        if localfile:
            try:
                self.book = xlrd.open_workbook(on_demand=True,
                                               filename=url)
            except:
                print("Error on %s" % url)
        else:
            try:
                conn = urlopen(url)
                
            except URLError as strerr:
                print("\nURL Error reading url: %s\n %s" % (url, strerr))
            
            except: 
                print("\nGeneral Error reading url: %s\n" % url)
                
            else:
                try:
                    self.book = xlrd.open_workbook(on_demand=True,
                                                   file_contents=conn.read())
                except:
                    print("Error on %s" % url)
            
            finally:
                conn.close()
            
        if self.book:
            self.datemode = self.book.datemode
                
    def xlCellValue(self, x):
        return xlValue(x, self.datemode, self.hash_comments)

    def sheet(self, sheet=None):
        '''
        Returns worksheet object by name or index, 
        
        Example:
        > wb = XLSReader('test.xls')
        > wb.sheet('sheet2')   # returns sheet named 'sheet2'
        > wb.sheet(3)          # returns sheet with index 3
        
        Default (if no argument is pased) returns sheet index = 0 
        
        '''
        if type(sheet) == int:
            sheet_name, sheet_index = None, sheet
        else:
            sheet_name, sheet_index = sheet, 0
        
        try:
            if sheet_name:
                self.sh = self.book.sheet_by_name(sheet)
            else:
                self.sh = self.book.sheet_by_index(sheet)
        except:
            print("Invalid sheet %s %s" % (sheet, type(sheet)))
            return None
        
        self.ncolumns = self.sh.ncols
        self.nrows = self.sh.nrows
        
        return self.sh
        
    def sheet_as_db(self, sheet=None, 
                    header=True, dkey=0,
                    startrow=0, numrows=None):
        '''
        Reads rows in a given sheet. Returns (refcolumn, hdr, qdata)
        
        sheet:  Name of sheet or 0-indexed sheet number.          
                        
        header: True--return rows as dict objects, with 'startrow' as keys.
                False--return rows as list
        
        dkey:   Column to serve as the dict key for qdata.
                '-1' uses the row number as key.
        
        startrow: Begin reading at this row (0 indexed), ignore ealier rows                
        
        numrows: number of rows to read, None=read all
        
        '''
        class sheetdb(object):
            pass
            
        self.sh = self.sheet(sheet)
        
        cleanrow_ = lambda row_: [x if x is not '' else None for x in row_]
            
        startloc = 1 if dkey == 0 else 0
        
        hdr = None
        if header:
            hdr = [get_xlvalue(h) for h in self.sh.row(startrow)]
            def rowValues(row_, loc): 
                return dict(zip(hdr[loc:],
                                cleanrow_(row_[loc:])))
                                                   
        else:
            rowValues = lambda row_, loc: cleanrow_(row_[loc:])
        
        qdata = []
        refcolumn = []
        
        if not numrows:
            numrows = self.sh.nrows
        else:
            numrows = startrow + numrows
            
        for xrow in range(startrow, numrows):
            try:
                xr = map(self.xlCellValue, self.sh.row(xrow))
                
            except:
                print("problem with row %s" % xrow)
                continue #skips a row if there's a problem
                
            else:
                xrvalues = rowValues(xr, startloc)
                qdata.append(xrvalues)
                if dkey >= 0:
                    refcolumn.append(xr[dkey])
                else:
                    refcolumn.append(xrow)
        
        nrows = len(qdata)
        
        if dkey >= 0:
            qdata = dict(zip(refcolumn, qdata))
        
        for attr in ['refcolumn', 'hdr', 'qdata']:
            setattr(sheetdb, attr, vars().get(attr, None))
            
        return sheetdb

    def read_sheet(self, sheet_name=None, sheet_index=0):
        def read_row(sheet, row):
            try:
                xr = map(self.xlCellValue, sh.row(row))
                
            except:
                #skips a row if there's a problem
                print("problem with row %s" % xrow)
                xr = []
                
            return xr
            
        sh = self.sheet(sheet_name, sheet_index)
        
        return [read_row(sh, xrow) for xrow in range(sh.nrows)]

class XLSXReader(object):
    def __init__(self, filename, hash_comments=True):
        self.book = None
        self.hash_comments = hash_comments
        
        # TODO: if we need to connect via url, try this kind of thing
        #  url = urlopen(filename)
        #  contents = url.read()
        #  stio = StringIO.StringIO(contents) 
        #  book = pyxl.reader.excel.load_workbook(filename = stio) 
        try:
            self.book = pyxl.reader.excel.load_workbook(filename=filename,
                                                        use_iterators=True)
        except:
            print("Error on %s" % filename)
            raise
                
    def xlCellValue(self, x): 
        if self.hash_comments and hasattr(x, "__iter__"):
            if x[0] == '#':
                return None
        try:
            return x.internal_value
        except:
            return x.value

    def sheet(self, sheet=None):
        '''
        Returns worksheet object by name or index, 
        
        Example:
        > wb = XLSXReader('test.xls')
        > wb.sheet('sheet2')   # returns sheet named 'sheet2'
        > wb.sheet(3)          # returns sheet with index 3
        
        Default (if no argument is pased) returns sheet index = 0 
        
        '''
        if type(sheet) == int:
            sheet_name, sheet_index = None, sheet
        else:
            sheet_name, sheet_index = sheet, 0
        
        try:
            if not sheet_name:
                sheet_name = self.book.get_sheet_names()[sheet_index]
        except:
            print("Invalid sheet: %s %s" % (sheet, type(sheet)))
            return None
        
        self.sh = self.book.get_sheet_by_name(sheet_name)
        
        return self.sh
        
    def sheet_as_db(self, sheet=None,
                    header=True, dkey=0,
                    startrow=0):
        '''
        Reads rows in a given sheet. Data is stored in 
        class member list self.qdata
        
        '''
        class sheetdb(object):
            pass
            
        self.sh = self.sheet(sheet)
        
        startloc = 1 if dkey == 0 else 0
        
        rowValues = lambda row_, loc: row_[loc:]
        hdr = None
        qdata = []
        refcolumn = []
        
        nrow = 0
        for row in self.sh.iter_rows():

            if nrow < startrow:
                nrow += 1
                continue

            if header and not hdr:
                hdr = [cell.internal_value for cell in row]
                def rowValues(row_, loc): 
                    return dict(zip(hdr[loc:],
                                    row_[loc:]))
                continue
            
            try:
                xr = map(self.xlCellValue, row)
                
            except:
                print("problem with row %s" % nrow)
                continue #skips a row if there's a problem
                
            else:
                xrvalues = rowValues(xr, startloc)
                qdata.append(xrvalues)
                if dkey >= 0:
                    refcolumn.append(xr[dkey])
                else:
                    refcolumn.append(nrow)
        
        if dkey >= 0:
            qdata = dict(zip(refcolumn, qdata))
                
        for attr in ['refcolumn', 'hdr', 'qdata']:
            setattr(sheetdb, attr, vars().get(attr, "what"))
            
        return sheetdb         

class XLdb(object):
    '''
    Container class for loading an Excel Database
    
    Class creates member dictionary with data from the specified 
    spreadsheet.
    
    startrow:       Begin reading at this row (0 indexed), ignore ealier rows
                    
    header:         True--return rows as dict objects, with 'startrow' as keys.
                    False--return rows as list
                    
    sheet:     Name of sheet or 0-indexed sheet number.          
    
    idx_column:     Column to serve as the dict key for qdata.
                    '-1' uses the row number as key.
                    
    numrows:        Number of rows to read.  'None'(default) to read all.
                    
    hash_comments:  True returns 'None' if cell string starts with "#', 
                    ala Bloomberg data errors
                    
    '''
    def __init__(self, filepath, sheet=0,   
                 header=True, idx_column=0, 
                 startrow=0, numrows=None,
                 hash_comments=1,
                 localfile=True):

        self.filepath = filepath
        self.xlsbook = XLSReader(filepath, localfile=localfile)
        self.book = self.xlsbook.book
        
        self.sh = self.xlsbook.sheet(sheet)
        
        sheetdb = self.xlsbook.sheet_as_db(sheet,
                                         header, idx_column,
                                         startrow, numrows)
        
        for attr in ['refcolumn', 'hdr', 'qdata']:
            setattr(self, attr, getattr(sheetdb, attr, None))    
        
        self.book.unload_sheet(self.sh.name)
        self.book.release_resources()
    
    def get(self, key, default=None):
        if self.qdata:
            return self.qdata.get(key, default)
        else:
            return default
            
    def __getitem__(self, key):
        if self.qdata:
            return self.qdata.get(key, None)
        else:
            return None
    
    def column(self, columnName, reduce=True):
        '''Returns a list for the given column. 
        
        columnName:  The header for the data in qdata.
        reduce:      [default=True], reduces the list to remove Null values
        
        Will retun an empty list if columnName is not a header in the 
        spreadsheet data, not and error.
        
        '''
        colList = [self.qdata[recx].get(columnName, None) 
                   for recx in self.refcolumn]
                   
        if reduce:
            return filter(lambda x: x is not None, colList)
        else:
            return colList
            

class XLOut(object):
    '''
    Creates a workbook object for writing.  Defines 'styles', a dict of cell formats,
    e.g. "date".
    
    >>> wkb = XLout(filename)
    >>> wkb.write( cell_value, cell_row, cell_column, sheet_index)
    >>> wkb.write( date_value, cell_row2, cell_column2, sheet_index, "date")
    >>> wkb.save()
    
    N.B. 'write' does NOT save (because save can be slow), so you must make sure
    to issue a save before closing, or periodically to avoid losing data.
    
    Utility functions:
    
    freezepanes:  Allows freezing panes on a specified sheet
    timeseries:   Writes a dictionary {date: value-tuple} to a sheet
    sheet_names:  Property returning list of sheet names
    add_sheet:    Adds a sheet, to the end of the current list of sheets
    
    Most functions allow flexibility in specify sheets, to choose by 
    sheet index or sheet name
    
    '''
    datestyle = xlwt.XFStyle()
    datestyle.num_format_str='MM/DD/YYYY'
    
    pctstyle = xlwt.XFStyle()
    pctstyle.num_format_str="0.000%"
    
    defaultstyle = xlwt.XFStyle()
    
    styles = {"date": datestyle, "pct": pctstyle}
    
    def __init__(self, fname=None, sheets=["Sheet1"], fdir=None,
                       overwrite_ok=False):
        '''
        Creates an excel spreadsheet.
        
        fname:  filename, if not provided a temporary filename is used. filename
                extension will be changed to .xls
        
        sheets: specify what sheets the workbook should have.
                    1. a list of sheet names
                    2. an integer specify how many sheets
                    3. a string specifying the name of a single sheet
                
                Use the add_sheet method to add additional sheets.
                
        fdir:   directory in which to create the file.  If not provided 'fname'
                is taken to be the full path, or if a temporary filename is created
                the file is placed in a default directory
                
           
        '''
        if not fname:
            fd, fpath = mkstemp(dir=fdir, suffix=".xls")
            os.close(fd)
            
        else:
            # make sure file ends with .xls extension
            fname = '.'.join((fname.split('.')[0], "xls"))
            fpath = os.path.join(fdir if fdir else '', fname)
            
        self.filename = fpath
        self.overwrite_ok = overwrite_ok
        
        self.wkb = xlwt.Workbook()                

        # trying to be really flexible in terms of specify sheets
        self.sheet = {}
        try:
            nsheets = len(sheets)
        except:
            try:
                nsheets = int(sheets)
                sheets = ["Sheet%s"%str(n) for n in range(1, nsheets+1)]
            except:
                sheets = [str(sheets)]
                
        for n in range(len(sheets)):
            sheetname = sheets[n]
            self.sheet[n] = self.wkb.add_sheet(sheetname, overwrite_ok)
        
        self.save()
    
    @property
    def sheet_names(self):
        return [self.sheet[n].name.encode() for n in self.sheet]
        
    def get_sheet_name(self, n):
        try:
            return self.sheet[n].name
        
        except:
            return None
            
    def select_sheet(self, sheet=0):
        '''
        Return object for sheet. 'sheet' can be either the index number or
        name of sheet.
        
        '''
        if isinstance(sheet, str):
            try:
                sheet = self.sheet_names.index(sheet)
            
            except:    
                return None
            
        try:
            ws = self.sheet[sheet] 
            return ws
            
        except:
            return None
    
    def add_sheet(self, sheet_name=None):
        'Adds a sheet to the workbook'
        
        if not sheet_name:
            sheet_name = "Sheet%s" % str(len(self.sheets)+1)
        
        self.sheets.append(sheet_name)
        self.sheet.append(self.wkb.add_sheet(sheet_name, overwrite_ok))
        
        return self.sheet[-1]
        
    def write(self, value, row, col, sheet=0, format=None):
        '''
        Writes to specified cell.  DOES NOT SAVE.
        
        Returns None if successfull.  sys.exc_info if not.
        
        '''
        try:
            if type(value) == timedelta:
                #timedelta does not play nicely with xlwt
                value = str(value)
                
            style = self.styles.get(format, self.defaultstyle)
            
            ws = self.select_sheet(sheet)
            ws.write(row, col, value, style)
            
        except:
            print("\nError writing: %s to cell: %s %s %s with format %s" %
                  (value, row, col, sheet, format))
            return sys.exc_info()
        
        return None
        
    def freezepanes(self, row_, col_, sheet=0):
        ws = self.select_sheet(sheet)
        
        ws.set_panes_frozen(True)
        ws.set_remove_splits(True)
        
        ws.set_horz_split_pos(row_+1)
        ws.set_vert_split_pos(col_+1)
        
    def timeseries(self, xdata, sheet=0, hdr=None):
        '''
        xdata:  a dict object, {date_key: value}, where the key is assumed to
                be a datetime.date object.  'value' is either a single value or 
                a tuple (value1, value2, value3)
        
        hdr:    column headings, if not provided ['data', 'value1', ...]
        
        '''
        date_keys = xdata.keys()
        date_keys.sort()
        
        # how will we access data?
        testdata = xdata[date_keys[0]] 
        if hasattr(testdata, "__iter__"):
            get_column_value = lambda colnum: value[colnum-1]
            
        else:
            get_column_value = lambda n: value

        # create header if not provided
        if not hdr:
            hdr = ['date']

            if hasattr(testdata, "__iter__"):
                hdr.extend(['value%s'%str(n) for n in range(1, len(testdata)+1)])
                
            else:
                hdr.append('value1')
        
        # write header row
        for ncol in range(len(hdr)):
            rc = self.write(hdr[ncol], 0, ncol, sheet)
            if rc:
                print("\nXLOut.timeseries problem writing header:\n%s\n" % rc)
                print("Sheet %s, row %s, column %\n" % (sheet, 0, ncol))
                return rc
                
        # write data rows
        for nrow, dt in enumerate(date_keys, start=1):
            self.write(dt, nrow, 0, sheet, format='date')
            
            value = xdata[dt]
            for ncol in range(1, len(hdr)):
                rc = self.write(get_column_value(ncol), nrow, ncol, sheet)   
                if rc:
                    print("\nXLOut.timeseries problem writing data:\n%s\n" % rc)
                    print("Sheet %s, row %s, column %\n" % (sheet, nrow, ncol))
                    return rc
                    
    def save(self):
        self.wkb.save(self.filename)


def timeseries_toXLS(xdata, fname=None, fdir=None, hdr=None):
    '''
    Writes Timeseries data to a spreadsheet.
    
    xdata:  a dict object, {date_key: value}, where the key is assumed to
            be a datetime.date object.  'value' is either a single value or 
            a tuple (value1, value2, value3)
    
    fname:  filename, if not provided a temporary filename is used. filename
            extension will be changed to .xls
    
    fdir:   directory in which to create the file.  If not provided 'fname'
            is taken to be the full path, or if a temporary filename is created
            the file is placed in a default directory
    
    hdr:    column headings, if not provided ['data', 'value1', ...]
    '''

    # use tempfile.mkstemp to create file name if not provided
    if not fname:
        fd, fpath = mkstemp(dir=fdir, suffix=".xls")
        os.close(fd)
    else:
        # make sure file ends with .xls extension
        fname = '.'.join((fname.split('.')[0], "xls"))
        fpath = os.path.join(fdir if fdir else '', fname)
        
    wkb = XLOut(fpath)
    
    date_keys = xdata.keys()
    date_keys.sort()
    
    # create header if not provided
    if not hdr:
        testdata = xdata[date_keys[0]]        
        hdr = ['date']
        try:
            hdr.extend(['value'+str(n) for n in range(1, len(testdata))+1])
            
        except:
            hdr.append('value1')
    
    # write header row
    for ncol in range(len(hdr)):
        wkb.write(hdr[ncol], 0, ncol, 0)
    
    # write data rows
    for nrow, dt in enumerate(date_keys, start=1):
        wkb.write(dt, nrow, 0, 0, format='date')
        
        for ncol in range(1, len(hdr)):
            value = xdata[dt]
            wkb.write(value, nrow, ncol, format=hdr[ncol][1])
            
    wkb.save()

    return fpath

